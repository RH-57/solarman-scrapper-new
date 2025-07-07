import asyncio
import pandas as pd
from playwright.async_api import async_playwright
from datetime import datetime, timezone
import os
import json
from dotenv import load_dotenv
from openpyxl import load_workbook
import mysql.connector
from mysql.connector import Error
from colorama import init, Fore

# Init colorama
init(autoreset=True)

# Load .env
load_dotenv()

INTERVAL_HOURS = int(os.getenv("INTERVAL_HOURS", 2))
SAVE_TO_MYSQL = os.getenv("SAVE_TO_MYSQL", "false").lower() == "true"
USERS_FILE = os.path.join("users", "users.json")

DB_CONFIG = {
    "host": os.getenv("DB_HOST"),
    "port": int(os.getenv("DB_PORT", 3306)),
    "user": os.getenv("DB_USER"),
    "password": os.getenv("DB_PASSWORD"),
    "database": os.getenv("DB_NAME")
}

# Load users
#with open("users.json", "r") as f:
    #USERS = json.load(f)

def log_inline(message):
    now = datetime.now().strftime("[%H:%M:%S]")
    print(f"\r{now} {message}", end='', flush=True)

def log(message):
    now = datetime.now().strftime("[%Y-%m-%d %H:%M:%S]")
    if "✅" in message:
        print(f"{now} {Fore.GREEN}{message}")
    elif "❌" in message or "⚠️" in message:
        print(f"{now} {Fore.RED}{message}")
    else:
        print(f"{now} {message}")

def save_to_excel(device_id, new_data):
    sheet_name = str(device_id)
    username = new_data["Username"].iloc[0]  # Ambil username dari DataFrame
    excel_dir = "excel"
    os.makedirs(excel_dir, exist_ok=True)

    user_excel_path = os.path.join(excel_dir, f"{username}.xlsx")

    try:
        if os.path.exists(user_excel_path):
            book = load_workbook(user_excel_path)
            if sheet_name in book.sheetnames:
                old_data = pd.read_excel(user_excel_path, sheet_name=sheet_name)
                combined = pd.concat([new_data, old_data], ignore_index=True)
            else:
                combined = new_data
            with pd.ExcelWriter(user_excel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                combined.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            with pd.ExcelWriter(user_excel_path, engine='openpyxl') as writer:
                new_data.to_excel(writer, sheet_name=sheet_name, index=False)
        log(f"✅ File Excel [{username}.xlsx] updated untuk device [{sheet_name}].")
    except Exception as e:
        log(f"❌ Gagal simpan Excel [{username}.xlsx]: {e}")


def save_to_mysql(data):
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        if conn.is_connected():
            cursor = conn.cursor()
            query = """
                INSERT INTO logs_meter (username, name, device_id, device_name, dpe, unit, timestamp)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """
            cursor.execute(query, (
                data["Username"], data["Name"], data["Device ID"], data["Device Name"],
                data["Daily Positive Energy"], data["Unit"], data["DateTime"]
            ))
            conn.commit()
            log(f"✅ Data berhasil disimpan ke MySQL.")
        else:
            log("❌ Gagal terhubung ke database.")
    except Error as e:
        log(f"❌ Error saat menyimpan ke MySQL: {e}")
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()

async def get_account_info(page):
    try:
        await page.goto("https://globalpro.solarmanpv.com/setting/home")
        await page.wait_for_load_state("networkidle")

        labels = page.locator("div.cSub")
        values = {}
        for i in range(await labels.count()):
            label = (await labels.nth(i).inner_text()).strip()
            if label in ["Username", "Name"]:
                value_el = labels.nth(i).locator("xpath=following-sibling::div[1]")
                await value_el.wait_for(state="visible", timeout=5000)
                value_text = await value_el.inner_text()
                values[label] = value_text.strip()

        username = values.get("Username", "Unknown")
        name = values.get("Name", "Unknown")
        log(f"👤 Username: {username}")
        log(f"💼 Name: {name}")
        return username, name
    except Exception as e:
        log(f"⚠️ Gagal ambil info akun: {e}")
        return "Unknown", "Unknown"

async def close_popup(page):
    try:
        await page.evaluate("document.querySelector('.guideMask')?.remove()")
    except:
        pass
    try:
        await page.click("text=I know", timeout=3000)
    except:
        pass

async def scrape_device(p, user, device_id):
    username_env = user["username"]
    password_env = user["password"]
    session_file = os.path.join("sessions", f"session_{username_env}.json")

    log(f"🔄 Mulai scraping device {device_id} untuk user {username_env}")
    browser = await p.chromium.launch(headless=False, channel="chrome")
    try:
        if os.path.exists(session_file):
            context = await browser.new_context(storage_state=session_file)
        else:
            context = await browser.new_context()
            page = await context.new_page()
            await page.goto("https://globalpro.solarmanpv.com/login")
            await page.click("text=International")
            await page.click("text=Confirm")
             # Pastikan klik tab "Username"
            try:
                await page.wait_for_selector('.tabBar div.afterButton >> text=Username', timeout=5000)
                await page.click('.tabBar div.afterButton >> text=Username')
                log("🟦 Tab 'Username' dipilih.")
            except Exception as e:
                log(f"⚠️ Gagal memilih tab 'Username': {e}")

            await page.fill('input[placeholder="Username"]', username_env)
            await page.fill('input[placeholder="Password"]', password_env)
            log("🔐 Silakan login manual jika ada captcha. Tekan ENTER jika sudah login.")
            input()
            await page.wait_for_timeout(3000)
            await context.storage_state(path=session_file)

        page = await context.new_page()
        username_real, name_real = await get_account_info(page)

        await page.goto(f"https://globalpro.solarmanpv.com/station/device?id={device_id}")
        await close_popup(page)

        try:
            await page.wait_for_selector('.guideMask button', timeout=5000)
            await page.click('.guideMask button')
            await page.wait_for_selector('.guideMask', state='detached', timeout=5000)
        except:
            try:
                await page.evaluate("document.querySelector('.guideMask')?.remove()")
            except:
                log("⚠️ Tidak bisa hapus guideMask")


        try:
            node_title = page.locator("span.node-title").first
            await node_title.wait_for(state="visible", timeout=10000)
            raw_text = await node_title.get_attribute("title")
            device_name = raw_text.split("(")[0].strip()
            log(f"🔧 Device Name: {device_name}")
        except:
            device_name = "Unknown"
            log("⚠️ Gagal ambil nama device.")

        try:
            await page.click('td.curP div[title="Meter"]')
            await page.wait_for_selector('div.panel:has(span.fsLv3:has-text("Total"))')
            total_panel = await page.query_selector('div.panel:has(span.fsLv3:has-text("Total"))')
            daily_el = await total_panel.query_selector('li:has-text("Daily Positive Energy")')
        except:
            log(f"❌ Gagal temukan DPE untuk device {device_id}")
            return

        full_text = await daily_el.inner_text()
        log(f"📊 Data ditemukan: {full_text}")
        value_text = full_text.split("：")[-1].strip()
        import re
        match = re.match(r"([\d\.\-]+)\s*(\w+)", value_text)
        value = float(match.group(1)) if match else value_text
        unit = match.group(2) if match else ""

        now_local = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        now_utc = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")

        data_df = pd.DataFrame([{
            "DateTime": now_local,
            "Username": username_real,
            "Name": name_real,
            "Device ID": device_id,
            "Device Name": device_name,
            "Daily Positive Energy": value,
            "Unit": unit
        }])

        log(f"📋 Menyimpan data device {device_id}...")
        save_to_excel(device_id, data_df)

        if SAVE_TO_MYSQL:
            save_to_mysql({
                "Username": username_real,
                "Name": name_real,
                "Device ID": device_id,
                "Device Name": device_name,
                "Daily Positive Energy": value,
                "Unit": unit,
                "DateTime": now_utc
            })
        else:
            log("ℹ️ Penyimpanan ke MySQL dinonaktifkan. Hanya simpan ke Excel.")
    finally:
        await browser.close()

async def countdown(seconds):
    for remaining in range(seconds, 0, -1):
        log_inline(f" ⏳ Menunggu {remaining:2} detik...")
        await asyncio.sleep(1)
    print()

async def main():
    async with async_playwright() as p:
        while True:
            # 🔁 Baca ulang file users.json setiap loop
            with open(USERS_FILE, "r") as f:
                USERS = json.load(f)

            for user in USERS:
                for device_id in user["device_id"]:
                    await scrape_device(p, user, device_id)
                    await countdown(15)

            log(f"\n🕒 Menunggu {INTERVAL_HOURS} jam untuk scraping berikutnya...\n")
            await asyncio.sleep(INTERVAL_HOURS * 3600)
if __name__ == "__main__":
    asyncio.run(main())
