import asyncio
import pandas as pd
from playwright.async_api import async_playwright
from datetime import datetime, timezone
import os
from dotenv import load_dotenv
from openpyxl import load_workbook
from colorama import init, Fore, Style
from db import get_connection 
import mariadb
# Inisialisasi colorama
init(autoreset=True)

# Load environment
load_dotenv()

USERNAME = os.getenv("USERNAME")
PASSWORD = os.getenv("PASSWORD")
DEVICE_IDS = [id.strip() for id in os.getenv("DEVICE_ID", "").split(",") if id.strip()]
SESSION_FILE = os.getenv("SESSION_FILE", "solarman-session.json")
DATA_FILE = os.getenv("DATA_FILE", "daily_positive_energy.xlsx")
INTERVAL_HOURS = int(os.getenv("INTERVAL_HOURS", 2))
SAVE_TO_MYSQL = os.getenv("SAVE_TO_MYSQL", "false").lower() == "true"


def log_inline(message):
    now = datetime.now().strftime("[%H:%M:%S]")
    print(f"\r{now} {message}", end='', flush=True)

def log(message):
    now = datetime.now().strftime("[%Y-%m-%d %H:%M:%S]")
    if "✅" in message:
        print(f"{now} {Fore.GREEN}{message}")
    elif "❌" in message or "⚠️" in message:
        print(f"{now} {Fore.RED}{message}")
    else:
        print(f"{now} {message}")

def save_to_excel(device_id, new_data):
    sheet_name = str(device_id)
    try:
        if os.path.exists(DATA_FILE):
            book = load_workbook(DATA_FILE)
            if sheet_name in book.sheetnames:
                old_data = pd.read_excel(DATA_FILE, sheet_name=sheet_name)
                combined = pd.concat([new_data, old_data], ignore_index=True)
            else:
                combined = new_data
            with pd.ExcelWriter(DATA_FILE, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                combined.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            with pd.ExcelWriter(DATA_FILE, engine='openpyxl') as writer:
                new_data.to_excel(writer, sheet_name=sheet_name, index=False)
        log(f"✅ File daily_positive_energy.xlsx - sheet [{sheet_name}], Done.")
    except Exception as e:
        log(f"❌ Gagal menyimpan ke Excel: {e}")

from db import get_connection  # Impor dari file db.py

def save_to_mysql(data):
    conn = get_connection()
    if not conn:
        log("❌ Gagal terhubung ke database.")
        return

    try:
        cursor = conn.cursor()
        query = """
            INSERT INTO logs_meter (username, name, device_id, device_name, dpe, unit, timestamp)
            VALUES (?, ?, ?, ?, ?, ?, ?)  -- Gunakan tanda tanya (?) bukan %s di mariadb
        """
        cursor.execute(query, (
            data["Username"],
            data["Name"],
            data["Device ID"],
            data["Device Name"],
            data["Daily Positive Energy"],
            data["Unit"],
            data["DateTime"]
        ))
        conn.commit()
        log("✅ Data berhasil disimpan ke MySQL.")
    except mariadb.Error as e:
        log(f"❌ Error saat menyimpan ke MySQL: {e}")
    finally:
        cursor.close()
        conn.close()


async def get_account_info(page):
    try:
        await page.goto("https://globalpro.solarmanpv.com/setting/home")
        await page.wait_for_load_state("networkidle")

        labels = page.locator("div.cSub")
        values = {}

        count = await labels.count()
        for i in range(count):
            label = (await labels.nth(i).inner_text()).strip()
            if label in ["Username", "Name"]:
                value_el = labels.nth(i).locator("xpath=following-sibling::div[1]")
                await value_el.wait_for(state="visible", timeout=5000)
                value_text = await value_el.inner_text()
                values[label] = value_text.strip()

        username = values.get("Username", "Unknown")
        name = values.get("Name", "Unknown")
        log(f"👤 Username: {username}")
        log(f"💼 Name: {name}")
        return username, name

    except Exception as e:
        log(f"⚠️ Gagal mengambil data akun: {e}")
        return "Unknown", "Unknown"

async def scrape_device(p, device_id):
    log(f"🔄 Mulai scraping Device ID: {device_id}")
    browser = await p.chromium.launch(headless=False, channel="chrome")
    try:
        if os.path.exists(SESSION_FILE):
            context = await browser.new_context(storage_state=SESSION_FILE)
        else:
            context = await browser.new_context()
            page = await context.new_page()
            await page.goto("https://globalpro.solarmanpv.com/login")
            await page.click('text=International')
            await page.click('text=Confirm')
            await page.fill('input[placeholder="Username"]', USERNAME)
            await page.fill('input[placeholder="Password"]', PASSWORD)
            log("🔐 Login manual di browser (captcha), tekan ENTER setelah selesai.")
            input()
            await page.wait_for_timeout(3000)
            await context.storage_state(path=SESSION_FILE)

        page = await context.new_page()
        username, name = await get_account_info(page)

        await page.goto(f"https://globalpro.solarmanpv.com/station/device?id={device_id}")

        try:
            await page.wait_for_selector('.guideMask button', timeout=5000)
            await page.click('.guideMask button')
            await page.wait_for_selector('.guideMask', state='detached', timeout=5000)
        except:
            try:
                await page.evaluate("document.querySelector('.guideMask')?.remove()")
            except:
                log("⚠️ Tidak bisa hapus guideMask")

        try:
            node_title = page.locator("span.node-title").first
            await node_title.wait_for(state="visible", timeout=10000)
            raw_system_text = await node_title.get_attribute("title")
            system_text = raw_system_text.split("(")[0].strip()
            log(f"🔧 Device Name: {system_text}")
        except Exception as e:
            log(f"⚠️ Gagal mengambil nama device: {e}")
            system_text = "Unknown"

        try:
            await page.wait_for_selector('td.curP div[title="Meter"]', timeout=10000)
            await page.click('td.curP div[title="Meter"]')
        except:
            log(f"⚠️ Data tidak ditemukan di device {device_id}, lanjut ke device berikutnya.")
            return

        try:
            await page.wait_for_selector('div.panel:has(span.fsLv3:has-text("Total"))', timeout=10000)
            total_panel = await page.query_selector('div.panel:has(span.fsLv3:has-text("Total"))')
            daily_el = await total_panel.query_selector('li:has-text("Daily Positive Energy")')
        except:
            log(f"❌ Panel atau elemen tidak ditemukan untuk device {device_id}")
            return

        if daily_el:
            full_text = await daily_el.inner_text()
            log(f"📊 Data ditemukan: {full_text}")
            value_text = full_text.split('：')[-1].strip()
            import re
            match = re.match(r"([\d\.\-]+)\s*(\w+)", value_text)
            if match:
                value = float(match.group(1))
                unit = match.group(2)
            else:
                value = value_text
                unit = ""

            # Timestamp lokal untuk Excel
            timestamp_excel = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # Timestamp UTC untuk MySQL
            timestamp_mysql = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")

            new_data = pd.DataFrame([{
                "DateTime": timestamp_excel,
                "Username": username,
                "Name": name,
                "Device ID": device_id,
                "Device Name": system_text,
                "Daily Positive Energy": value,
                "Unit": unit,
            }])

            log(f"📋 Menyimpan data untuk device {device_id}...")
            save_to_excel(device_id, new_data)

            if SAVE_TO_MYSQL:
                save_to_mysql({
                    "Username": username,
                    "Name": name,
                    "Device ID": device_id,
                    "Device Name": system_text,
                    "Daily Positive Energy": value,
                    "Unit": unit,
                    "DateTime": timestamp_mysql  # UTC
                })
            else:
                log("ℹ️ Penyimpanan ke MySQL dinonaktifkan. Hanya simpan ke Excel.")

        else:
            log(f"❌ Daily Positive Energy tidak ditemukan untuk device {device_id}")
    finally:
        await browser.close()

async def countdown(seconds):
    for remaining in range(seconds, 0, -1):
        log_inline(f" ⏳ Menunggu {remaining:2} detik...")
        await asyncio.sleep(1)
    log("\n⏳Jeda selesai. Lanjutkan scraping.")

async def main():
    async with async_playwright() as p:
        while True:
            for device_id in DEVICE_IDS:
                await scrape_device(p, device_id)
                await countdown(15)
            log(f"\n🕒 Tunggu {INTERVAL_HOURS} jam untuk scraping selanjutnya...\n")
            await asyncio.sleep(INTERVAL_HOURS * 3600)

if __name__ == "__main__":
    asyncio.run(main())
